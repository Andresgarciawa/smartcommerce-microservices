from openpyxl import load_workbook
wb = load_workbook('Libro1-pruebas.xlsx')
sheet = wb.active
print('max_row', sheet.max_row)
for row in sheet.iter_rows(values_only=True):
    print(row)
